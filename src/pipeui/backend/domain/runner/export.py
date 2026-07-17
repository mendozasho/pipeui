"""Result-export builders — slice runner-execution/5.

Two consumption exports for the Results screen (PRD: Results surfacing — three tiers):

  build_results_report(run_result, *, source_name=None) -> {"columns": [...], "rows": [...]}
      The **transposed results report**: one row per function run (function × source),
      with pass/fail + metadata columns. INCLUDES runs that fully passed — there is no
      filtering on rows_failed, so the report is a complete record of every run.
      Accepts any runner output shape (all entry points collapse to one row list):
        - source-tied: run_pipeline(conn, source_id, run_type)
            -> { run_type, steps: [...] }   (steps carry function identity; the caller
               supplies source_name since steps do not carry it)
        - Functions-page cross-source: run_validation_across_sources(conn, function_id)
            -> { function_id, function_name, sources: [...] }  (each per-source entry IS
               a (function, source) RunResult carrying function identity + source_name)
        - set run: run_set_across_sources(conn, set_id)
            -> { set_id, set_name, sources: [ {source_name, steps: [...]}, ... ] }
               (flattened sources × steps; a source-level crash with no steps still
               yields one failed row so the report stays a complete record)

  write_results_csv / write_results_xlsx (report, dest_path) -> row count
      File writers for the results report (server-owned validation export). The report
      is one row per function run — always tiny — so these use stdlib csv / openpyxl
      directly rather than the DuckDB COPY path reserved for data tables.

  build_transformed_report(conn, source_id) -> {"columns": [...], "rows": [...]}
      The **transformed report**: the source's transformed data table after all assigned
      transforms completed (the latest staging table). A validation-only run writes no
      staging table, so this returns an empty payload rather than raising — that is the
      #193 staging-export-failure fix for a mixed validation/transform set.

File-download exports (#110 — large tables must never round-trip through JSON):

  get_staging_meta(conn, source_id) -> {"exists", "row_count", "columns"}
      Cheap preflight for the download flow — no row materialization.

  write_transformed_csv(conn, source_id, dest_path) -> row count | None
      DuckDB-native COPY TO; None when no staging table exists.

  write_transformed_xlsx(conn, source_id, dest_path) -> row count | None
      openpyxl write-only streaming over Arrow batches; raises ValueError when the
      table exceeds the xlsx sheet row limit. None when no staging table exists.

§14: this is a workflow-layer module; api/ route modules call it, never schema/ directly.
"""
from __future__ import annotations

import uuid
from typing import Any

import duckdb

from pipeui.backend.data.runner.staging import latest_staging
from pipeui.backend.domain.runner.run import get_staging_rows

# xlsx sheet limit is 1,048,576 rows; one is reserved for the header.
XLSX_MAX_DATA_ROWS = 1_048_575


# Columns surfaced in the transposed results report, in order — one row per function
# run (function × source). Every row carries these keys (missing values default to
# None) so the exported file is rectangular.
_RESULT_COLUMNS = [
    "function_name",
    "function_type",
    "source_name",
    "status",
    "rows_passed",
    "rows_failed",
    "pass_rate",
    "error",
    "result_id",
]


def _result_row(entry: dict[str, Any], *, source_name: str | None = None) -> dict[str, Any]:
    """Project one runner result entry onto the transposed report's column set.

    `source_name` fills the column when the entry itself does not carry it (source-tied
    steps and set-run nested steps get it from their surrounding context).
    """
    row: dict[str, Any] = {col: entry.get(col) for col in _RESULT_COLUMNS}
    if row.get("source_name") is None:
        row["source_name"] = source_name
    return row


def build_results_report(
    run_result: dict | None,
    *,
    source_name: str | None = None,
) -> dict[str, Any]:
    """Build the transposed results report from any runner output shape.

    Returns {"columns": [...], "rows": [...]} — one row per function run
    (function × source), including runs that fully passed. Returns an empty
    payload for a None/empty run result.
    """
    if not run_result:
        return {"columns": list(_RESULT_COLUMNS), "rows": []}

    # Source-tied shape: a flat `steps` list; the steps carry function identity but
    # not the source, which the caller passes in.
    if "steps" in run_result:
        rows = [_result_row(s, source_name=source_name) for s in run_result["steps"] or []]
        return {"columns": list(_RESULT_COLUMNS), "rows": rows}

    rows = []
    for entry in run_result.get("sources") or []:
        entry_source = entry.get("source_name")
        if "steps" in entry:
            # Set shape: per-source entries hold nested per-function steps.
            steps = entry.get("steps") or []
            for step in steps:
                rows.append(_result_row(step, source_name=entry_source))
            if not steps and entry.get("error"):
                # A source-level crash ran no functions; it must still appear as a
                # failure so the report stays a complete record of the run.
                rows.append(_result_row(
                    {"status": "failed", "error": entry.get("error")},
                    source_name=entry_source,
                ))
        else:
            # Cross-source shape: each entry IS one (function, source) RunResult.
            rows.append(_result_row(entry, source_name=entry_source))
    return {"columns": list(_RESULT_COLUMNS), "rows": rows}


def _format_report_cell(col: str, value: Any) -> Any:
    if col == "pass_rate" and isinstance(value, (int, float)):
        return f"{value * 100:.1f}%"
    return value


def write_results_csv(report: dict[str, Any], dest_path: str) -> int:
    """Write a results report ({"columns", "rows"}) to dest_path as CSV; return the
    data row count. None cells become empty CSV fields."""
    import csv

    columns = report["columns"]
    with open(dest_path, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=columns)
        writer.writeheader()
        for row in report["rows"]:
            writer.writerow({c: _format_report_cell(c, row.get(c)) for c in columns})
    return len(report["rows"])


def write_results_xlsx(report: dict[str, Any], dest_path: str) -> int:
    """Write a results report ({"columns", "rows"}) to dest_path as xlsx; return the
    data row count."""
    from openpyxl import Workbook

    columns = report["columns"]
    wb = Workbook(write_only=True)
    ws = wb.create_sheet("Results")
    ws.append(columns)
    for row in report["rows"]:
        ws.append([_xlsx_safe(_format_report_cell(c, row.get(c))) for c in columns])
    wb.save(dest_path)
    return len(report["rows"])


def build_transformed_report(
    conn: duckdb.DuckDBPyConnection,
    source_id: uuid.UUID,
) -> dict[str, Any]:
    """Build the transformed report (the transformed data table) for a source.

    Returns the latest staging table as {"columns": [...], "rows": [...]} — the
    transformed data after all assigned transforms completed. Returns an empty payload
    (not an error) when no transform has run, so a mixed validation/transform set
    exports cleanly (#193).
    """
    return get_staging_rows(conn, source_id)


def get_staging_meta(
    conn: duckdb.DuckDBPyConnection,
    source_id: uuid.UUID,
) -> dict[str, Any]:
    """Return {"exists", "row_count", "columns"} for the source's latest staging table.

    Never materializes rows — one count(*) plus a DESCRIBE. {"exists": False,
    "row_count": 0, "columns": []} when no transform has run yet.
    """
    staged = latest_staging(conn, source_id)
    if staged is None:
        return {"exists": False, "row_count": 0, "columns": []}
    tname, _ts = staged
    (row_count,) = conn.execute(f'SELECT count(*) FROM "{tname}"').fetchone()
    columns = [r[0] for r in conn.execute(f'DESCRIBE "{tname}"').fetchall()]
    return {"exists": True, "row_count": int(row_count), "columns": columns}


def write_transformed_csv(
    conn: duckdb.DuckDBPyConnection,
    source_id: uuid.UUID,
    dest_path: str,
) -> int | None:
    """Write the latest staging table to dest_path as CSV; return the row count.

    Returns None when no staging table exists (mirrors the #193 empty-payload
    contract). NULL/NaN cells become empty CSV fields natively — no JSON
    round-trip, no Python-side row handling.
    """
    staged = latest_staging(conn, source_id)
    if staged is None:
        return None
    tname, _ts = staged
    # COPY TO cannot take a bound parameter for the path; the staging table name
    # is repo-generated (staging_{hex}_{ts}) so only the path needs escaping.
    escaped = dest_path.replace("'", "''")
    row = conn.execute(
        f'COPY (SELECT * FROM "{tname}") TO \'{escaped}\' (FORMAT CSV, HEADER)'
    ).fetchone()
    return int(row[0])


def write_transformed_xlsx(
    conn: duckdb.DuckDBPyConnection,
    source_id: uuid.UUID,
    dest_path: str,
) -> int | None:
    """Write the latest staging table to dest_path as xlsx; return the row count.

    Returns None when no staging table exists. Raises ValueError before writing
    anything when the table exceeds XLSX_MAX_DATA_ROWS — the format cannot hold
    it and the caller should steer the user to CSV. Streams Arrow record batches
    through an openpyxl write-only workbook so the full table is never
    materialized in memory.
    """
    staged = latest_staging(conn, source_id)
    if staged is None:
        return None
    tname, _ts = staged

    (row_count,) = conn.execute(f'SELECT count(*) FROM "{tname}"').fetchone()
    row_count = int(row_count)
    if row_count > XLSX_MAX_DATA_ROWS:
        raise ValueError(
            f"Table has {row_count:,} rows — the xlsx format holds at most "
            f"{XLSX_MAX_DATA_ROWS:,} data rows. Export as CSV instead."
        )

    from openpyxl import Workbook

    columns = [r[0] for r in conn.execute(f'DESCRIBE "{tname}"').fetchall()]
    wb = Workbook(write_only=True)
    ws = wb.create_sheet("Transformed")
    ws.append(columns)

    reader = conn.execute(f'SELECT * FROM "{tname}"').to_arrow_reader(10_000)
    for batch in reader:
        for row in batch.to_pylist():
            ws.append([_xlsx_safe(row[c]) for c in columns])
    wb.save(dest_path)
    return row_count


def _xlsx_safe(value: Any) -> Any:
    """Coerce a cell value to a type openpyxl accepts natively."""
    if isinstance(value, (bytes, bytearray)):
        return value.decode("utf-8", errors="replace")
    if isinstance(value, (list, dict)):
        return str(value)
    if isinstance(value, float) and value != value:  # NaN is not representable in xlsx
        return None
    return value
