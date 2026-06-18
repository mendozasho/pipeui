from __future__ import annotations

import csv
import datetime
import os
import re
import uuid
from pathlib import Path

import duckdb

from pipeui.backend.data.base.ids import content_hash_id
from pipeui.backend.data.base.schema.constants import IngestionMethod, DUCKDB_TO_PYTHON
from pipeui.backend.data.base.db import infer_column_types, get_db_path
from pipeui.helpers import infer_pattern
from pipeui.backend.data.base.fails import FailedRegistryEntry
from pipeui.backend.data.sources.registry import SourceRegistryEntry, SourceRegistryUpdate
from pipeui.backend.data.sources.columns import ColumnRegistryEntry


class CreateFlowCache:
    """The create flow's column-type staging cache (the ``_stage_create_flow`` temp
    table): holds confirmed column names/types/PK while a source is being created,
    before it is committed. Part of the create flow (this module) — distinct from the
    runner's transformed-output staging store in ``staging.py`` (they only share the
    word "staging")."""

    def __init__(self, conn: duckdb.DuckDBPyConnection) -> None:
        self._conn = conn
        conn.execute("""
            CREATE TEMP TABLE IF NOT EXISTS _stage_create_flow (
                column_name    VARCHAR NOT NULL,
                column_type    VARCHAR NOT NULL,
                is_primary_key BOOLEAN NOT NULL DEFAULT false
            )
        """)

    def stage_columns(self, columns: list[tuple[str, str]]) -> None:
        """Replace any staged columns with the given (name, type) pairs."""
        self._conn.execute("DELETE FROM _stage_create_flow")
        self._conn.executemany(
            "INSERT INTO _stage_create_flow (column_name, column_type) VALUES (?, ?)",
            columns,
        )

    def set_primary_key(self, column_name: str) -> None:
        """Mark exactly one staged column as the primary key (clears the others)."""
        self._conn.execute("UPDATE _stage_create_flow SET is_primary_key = false")
        self._conn.execute(
            "UPDATE _stage_create_flow SET is_primary_key = true WHERE column_name = ?",
            [column_name],
        )

    def get_staged(self) -> list[dict]:
        """Return staged columns as dicts (column_name, column_type, is_primary_key)."""
        rows = self._conn.execute(
            "SELECT column_name, column_type, is_primary_key FROM _stage_create_flow"
        ).fetchall()
        return [
            {"column_name": r[0], "column_type": r[1], "is_primary_key": r[2]}
            for r in rows
        ]

    def get_primary_key(self) -> str | None:
        """Return the staged primary-key column name, or None."""
        row = self._conn.execute(
            "SELECT column_name FROM _stage_create_flow WHERE is_primary_key = true LIMIT 1"
        ).fetchone()
        return row[0] if row else None

    def clear(self) -> None:
        """Remove all staged columns."""
        self._conn.execute("DELETE FROM _stage_create_flow")

    def drop(self) -> None:
        """Drop the staging temp table if it exists."""
        self._conn.execute("DROP TABLE IF EXISTS _stage_create_flow")


def find_source_by_pattern(
    conn: duckdb.DuckDBPyConnection,
    filename: str,
) -> uuid.UUID | None:
    """Return the source_id of an existing source whose pattern matches filename's stem, or None."""
    stem = Path(filename).stem
    rows = conn.execute(
        "SELECT source_id, pattern FROM source_registry WHERE pattern IS NOT NULL"
    ).fetchall()
    for source_id, pattern in rows:
        try:
            if re.fullmatch(pattern, stem):
                return uuid.UUID(str(source_id))
        except re.error:
            continue
    return None


def peek_header_columns(file_path: str) -> list[str]:
    """Return only the header (first row) column names from a CSV/TSV or XLSX file.

    Reads just the first row — for XLSX via openpyxl's read-only iterator, for
    delimited files via a single ``csv.reader`` row — so a file with hundreds of
    thousands of rows never has to be materialized to list its columns. Blank /
    whitespace-only header cells are dropped. Returns ``[]`` for an empty file or an
    unsupported extension; the caller treats ``[]`` as "couldn't detect" and falls
    back to a free-text column input.
    """
    ext = Path(file_path).suffix.lower()
    if ext == ".xlsx":
        import openpyxl  # lazy import — only needed on the upload path
        wb = openpyxl.load_workbook(file_path, read_only=True)
        try:
            ws = wb[wb.sheetnames[0]]
            header = next(ws.iter_rows(values_only=True), ())
        finally:
            wb.close()  # read-only workbooks must be closed to release the file
        return [str(c).strip() for c in header if c is not None and str(c).strip()]
    if ext in (".csv", ".tsv"):
        delimiter = "\t" if ext == ".tsv" else ","
        with open(file_path, newline="", encoding="utf-8-sig") as fh:
            row = next(csv.reader(fh, delimiter=delimiter), [])
        return [c.strip().strip("\"'") for c in row if c and c.strip()]
    return []


def create_source(
    conn: duckdb.DuckDBPyConnection,
    file_path: str,
    source_name: str,
    primary_key: str,
    ingestion_method: str = "upsert",
) -> tuple[uuid.UUID | None, FailedRegistryEntry]:
    """Execute the source-create flow as one atomic transaction."""
    failed = FailedRegistryEntry()

    if not IngestionMethod.accepted(ingestion_method):
        failed.add(None, f"Invalid ingestion method: {ingestion_method!r}")
        return None, failed

    # Try to infer the pattern of the filename for future searches
    pattern = infer_pattern(Path(file_path).name)

    # Try to infer column types from the file. Will later add the ones not in the column registry
    raw_columns = infer_column_types(conn, file_path)
    # Defensive fallback: any type not in our known set (e.g. from a mock or future DuckDB type)
    columns = [(name, t if t in DUCKDB_TO_PYTHON else "VARCHAR") for name, t in raw_columns]

    # [1] [stage in create-flow cache]
    # We stage so that any errors can be rolled back to previous working state.
    cache = CreateFlowCache(conn)
    cache.stage_columns(columns)
    cache.set_primary_key(primary_key)  # [DEFERRED] PK uniqueness not validated (§6, CLAUDE.md M4)

    # The registry table requires the date we ingested the file, so we can keep track.
    # Future builds may use this as a way of rolling back to a previous version.
    date_ingested = datetime.datetime.fromtimestamp(os.path.getmtime(file_path))

    # [2] [build and validate SourceRegistryEntry]
    # We build out the entry with a pydantic-based class to ensure the data is validated and correct.
    try:
        entry = SourceRegistryEntry(
            source_name=source_name,
            date_ingested=date_ingested,
            ingestion_method=ingestion_method,
            pattern=pattern,
            primary_key=primary_key,
        )
    except Exception as exc:
        failed.add(None, str(exc))  # should be an error from the pydantic validation
        return None, failed

    # TODO: This should likely be added as part of the original step.
    #  We aren't creating the path based off the variables
    db_path = get_db_path(conn)
    entry.generate_table_url(db_path)

    staged = cache.get_staged()  # used for the column table

    # one transaction covers source_registry + column_registry + source_column_map
    conn.execute("BEGIN")
    try:
        conn.execute(
            """
            INSERT INTO source_registry
                (source_id, content_hash_id, source_name, date_ingested,
                 date_registered, ingestion_method, pattern, primary_key, table_url)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            [
                entry.source_id,
                entry.content_hash_id,
                entry.source_name,
                entry.date_ingested,
                entry.date_registered,
                entry.ingestion_method,
                entry.pattern,
                entry.primary_key,
                entry.table_url,
            ],
        )

        for col in staged:
            col_entry = ColumnRegistryEntry(
                column_name=col["column_name"],
                column_type=col["column_type"],
            )
            # column_registry is shared: same (name, type) → same content_hash_id.
            # Reuse the existing column_id if this definition is already registered.
            existing_col = conn.execute(
                "SELECT column_id FROM column_registry WHERE content_hash_id = ?",
                [col_entry.content_hash_id],
            ).fetchone()
            if existing_col:
                col_entry.column_id = existing_col[0]
            else:
                conn.execute(
                    """
                    INSERT INTO column_registry
                        (column_id, content_hash_id, column_name, column_type)
                    VALUES (?, ?, ?, ?)
                    """,
                    [
                        col_entry.column_id,
                        col_entry.content_hash_id,
                        col_entry.column_name,
                        col_entry.column_type,
                    ],
                )

            # Map row written directly — no pydantic object (CLAUDE.md Architecture)
            map_id = content_hash_id(
                "source_column_map", str(entry.source_id), str(col_entry.column_id)
            )
            conn.execute(
                "INSERT INTO source_column_map (source_column_map_id, column_id, source_id) VALUES (?, ?, ?)",
                [map_id, col_entry.column_id, entry.source_id],
            )

        conn.execute("COMMIT")
        return entry.source_id, failed

    except Exception as exc:
        conn.execute("ROLLBACK")
        failed.add(entry, str(exc))
        return None, failed


def update_source(
    conn: duckdb.DuckDBPyConnection,
    source_id: uuid.UUID,
    **updates,
) -> tuple[uuid.UUID | None, FailedRegistryEntry]:
    """Update a source registry row, rejecting edits that collide on content_hash_id."""
    failed = FailedRegistryEntry()

    row = conn.execute(
        """SELECT source_id, source_name, date_ingested, date_registered,
                  ingestion_method, pattern, primary_key, table_url
           FROM source_registry WHERE source_id = ?""",
        [source_id],
    ).fetchone()

    if row is None:
        failed.add(None, f"source_id {source_id!r} not found")
        return None, failed

    existing = SourceRegistryEntry(
        source_id=row[0],
        source_name=row[1],
        date_ingested=row[2],
        date_registered=row[3],
        ingestion_method=row[4],
        pattern=row[5],
        primary_key=row[6],
        table_url=row[7],
    )

    update = SourceRegistryUpdate.from_existing(existing, **updates)

    # Collision check: new hash must not already exist on a different row (Principle 1)
    collision = conn.execute(
        "SELECT 1 FROM source_registry WHERE content_hash_id = ? AND source_id != ?",
        [update.content_hash_id, source_id],
    ).fetchone()

    if collision:
        failed.add(existing, "content_hash_id collision")
        return None, failed

    set_fields: dict[str, object] = {k: v for k, v in updates.items()}
    set_fields["content_hash_id"] = update.content_hash_id
    set_clause = ", ".join(f"{k} = ?" for k in set_fields)
    values = list(set_fields.values()) + [source_id]

    conn.execute("BEGIN")
    try:
        conn.execute(
            f"UPDATE source_registry SET {set_clause} WHERE source_id = ?",
            values,
        )
        conn.execute("COMMIT")
        return source_id, failed
    except Exception as exc:
        conn.execute("ROLLBACK")
        failed.add(existing, str(exc))
        return None, failed
