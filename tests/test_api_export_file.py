"""Behavioral guarantees for the transform file-download flow (#110, §13).

GET /pipelines/{source_id}/export/transformed/file?format=csv|xlsx
  1. Streams the latest staging table as an attachment (Content-Disposition)
     with the {source}_{date}_transform.{ext} filename — no JSON round-trip.
  2. NULL/NaN staging cells export as empty fields, never a 500.
  3. xlsx round-trips through openpyxl; a table over the xlsx sheet row limit
     is rejected with 422 before anything is written.
  4. 404 for an unknown source and when no staging table exists yet;
     422 for a malformed source_id or an unsupported format.
  5. The temp file is always removed — after a successful download and on
     every error branch.

GET /pipelines/{source_id}/staging/meta
  6. {"exists": False, "row_count": 0, "columns": []} before any transform;
     exists/row_count/columns after; 404 for an unknown source.
"""
from __future__ import annotations

import csv
import io
import os
import tempfile
import time
import uuid

import pandas as pd
import pytest
from fastapi import FastAPI
from fastapi.testclient import TestClient

import pipeui.backend.domain.runner.export as export_mod
from pipeui.middleware.pipelines import router, get_conn
from pipeui.backend.domain.sources.create import create_source
from pipeui.backend.domain.sources.ingestion import ingest_source
from pipeui.backend.data.runner.staging import write_staging_table


@pytest.fixture
def client(db):
    app = FastAPI()
    app.include_router(router)
    app.dependency_overrides[get_conn] = lambda: db
    yield TestClient(app)


@pytest.fixture
def tmpfile_spy(monkeypatch):
    """Record every path handed out by tempfile.mkstemp so cleanup is provable."""
    created: list[str] = []
    real_mkstemp = tempfile.mkstemp

    def spying_mkstemp(*args, **kwargs):
        fd, path = real_mkstemp(*args, **kwargs)
        created.append(path)
        return fd, path

    monkeypatch.setattr(tempfile, "mkstemp", spying_mkstemp)
    return created


def _make_csv(tmp_path, name, columns, rows):
    p = tmp_path / name
    with open(p, "w", newline="") as f:
        w = csv.writer(f)
        w.writerow(columns)
        w.writerows(rows)
    return str(p)


def _register_and_ingest(db, tmp_path):
    csv_path = _make_csv(
        tmp_path, "src.csv",
        ["id", "val"],
        [["1", "10"], ["2", "20"], ["3", "30"]],
    )
    source_id, _ = create_source(
        conn=db,
        file_path=csv_path,
        source_name="export_file_test",
        primary_key="id",
        ingestion_method="upsert",
    )
    ingest_source(conn=db, source_id=source_id, file_path=csv_path)
    return source_id


def _stage(db, source_id, df=None):
    if df is None:
        df = pd.DataFrame({"id": [1, 2, 3], "val": [10, 20, 30], "doubled": [20, 40, 60]})
    write_staging_table(db, source_id, df, int(time.time()))
    return df


# ---------------------------------------------------------------------------
# CSV download
# ---------------------------------------------------------------------------

@pytest.mark.integration
def test_csv_download_streams_attachment_with_transform_filename(client, db, tmp_path):
    """Guarantee 1: the CSV download is an attachment carrying the staging table."""
    source_id = _register_and_ingest(db, tmp_path)
    _stage(db, source_id)

    resp = client.get(f"/pipelines/{source_id}/export/transformed/file?format=csv")
    assert resp.status_code == 200, resp.text
    disposition = resp.headers["content-disposition"]
    assert "attachment" in disposition
    assert "export_file_test_" in disposition
    assert "_transform.csv" in disposition
    assert resp.headers["content-type"].startswith("text/csv")

    rows = list(csv.DictReader(io.StringIO(resp.text)))
    assert len(rows) == 3
    assert sorted(int(r["doubled"]) for r in rows) == [20, 40, 60]


@pytest.mark.integration
def test_csv_download_writes_nulls_as_empty_fields(client, db, tmp_path):
    """Guarantee 2: NaN/None staging cells become empty CSV fields, not a 500 (#262 data)."""
    source_id = _register_and_ingest(db, tmp_path)
    _stage(db, source_id, pd.DataFrame({
        "id": [1, 2, 3],
        "monthly_spend": [10.5, float("nan"), 30.0],
        "name": ["a", None, "c"],
    }))

    resp = client.get(f"/pipelines/{source_id}/export/transformed/file?format=csv")
    assert resp.status_code == 200, resp.text
    by_id = {r["id"]: r for r in csv.DictReader(io.StringIO(resp.text))}
    assert by_id["2"]["monthly_spend"] == ""
    assert by_id["2"]["name"] == ""
    assert by_id["1"]["monthly_spend"] == "10.5"


# ---------------------------------------------------------------------------
# xlsx download
# ---------------------------------------------------------------------------

@pytest.mark.integration
def test_xlsx_download_round_trips_through_openpyxl(client, db, tmp_path):
    """Guarantee 3: the xlsx download opens in openpyxl and carries the staging rows."""
    from openpyxl import load_workbook

    source_id = _register_and_ingest(db, tmp_path)
    _stage(db, source_id)

    resp = client.get(f"/pipelines/{source_id}/export/transformed/file?format=xlsx")
    assert resp.status_code == 200, resp.text
    assert "attachment" in resp.headers["content-disposition"]
    assert "_transform.xlsx" in resp.headers["content-disposition"]

    wb = load_workbook(io.BytesIO(resp.content))
    ws = wb.active
    grid = [[c.value for c in row] for row in ws.iter_rows()]
    assert grid[0] == ["id", "val", "doubled"]
    assert sorted(r[2] for r in grid[1:]) == [20, 40, 60]


@pytest.mark.integration
def test_xlsx_download_rejects_table_over_sheet_row_limit(client, db, tmp_path, monkeypatch, tmpfile_spy):
    """Guarantee 3: a table too large for the xlsx format is a 422 steering to CSV."""
    source_id = _register_and_ingest(db, tmp_path)
    _stage(db, source_id)
    monkeypatch.setattr(export_mod, "XLSX_MAX_DATA_ROWS", 2)

    resp = client.get(f"/pipelines/{source_id}/export/transformed/file?format=xlsx")
    assert resp.status_code == 422, resp.text
    assert "CSV" in resp.json()["detail"]
    assert tmpfile_spy and not os.path.exists(tmpfile_spy[-1]), "temp file leaked on the 422 branch"


# ---------------------------------------------------------------------------
# Error branches
# ---------------------------------------------------------------------------

@pytest.mark.integration
def test_download_error_statuses(client, db, tmp_path):
    """Guarantee 4: 404 unknown source / no staging; 422 malformed id / bad format."""
    assert client.get(f"/pipelines/{uuid.uuid4()}/export/transformed/file").status_code == 404
    assert client.get("/pipelines/null/export/transformed/file").status_code == 422

    source_id = _register_and_ingest(db, tmp_path)
    resp = client.get(f"/pipelines/{source_id}/export/transformed/file?format=pdf")
    assert resp.status_code == 422
    assert "pdf" in resp.json()["detail"]

    resp = client.get(f"/pipelines/{source_id}/export/transformed/file")
    assert resp.status_code == 404, "no staging table yet must 404, not stream an empty file"
    assert "No transformed data" in resp.json()["detail"]


# ---------------------------------------------------------------------------
# Temp-file lifecycle
# ---------------------------------------------------------------------------

@pytest.mark.integration
def test_temp_file_removed_after_successful_download(client, db, tmp_path, tmpfile_spy):
    """Guarantee 5: the background task removes the temp file once the body is sent."""
    source_id = _register_and_ingest(db, tmp_path)
    _stage(db, source_id)

    resp = client.get(f"/pipelines/{source_id}/export/transformed/file?format=csv")
    assert resp.status_code == 200
    assert tmpfile_spy, "route did not allocate its temp file via tempfile.mkstemp"
    # TestClient runs BackgroundTasks before returning the response.
    assert not os.path.exists(tmpfile_spy[-1]), "temp file leaked after a successful download"


@pytest.mark.integration
def test_temp_file_removed_when_no_staging_exists(client, db, tmp_path, tmpfile_spy):
    """Guarantee 5: the 404-no-staging branch also removes its temp file."""
    source_id = _register_and_ingest(db, tmp_path)

    resp = client.get(f"/pipelines/{source_id}/export/transformed/file?format=csv")
    assert resp.status_code == 404
    assert tmpfile_spy and not os.path.exists(tmpfile_spy[-1]), "temp file leaked on the 404 branch"


# ---------------------------------------------------------------------------
# Staging meta preflight
# ---------------------------------------------------------------------------

@pytest.mark.integration
def test_staging_meta_before_and_after_transform(client, db, tmp_path):
    """Guarantee 6: meta reports exists/row_count/columns without materializing rows."""
    source_id = _register_and_ingest(db, tmp_path)

    resp = client.get(f"/pipelines/{source_id}/staging/meta")
    assert resp.status_code == 200, resp.text
    assert resp.json() == {"exists": False, "row_count": 0, "columns": []}

    _stage(db, source_id)
    resp = client.get(f"/pipelines/{source_id}/staging/meta")
    assert resp.status_code == 200, resp.text
    data = resp.json()
    assert data["exists"] is True
    assert data["row_count"] == 3
    assert set(data["columns"]) == {"id", "val", "doubled"}


@pytest.mark.integration
def test_staging_meta_404_for_unknown_source(client, db):
    """Guarantee 6: meta 404s for an unknown source_id."""
    assert client.get(f"/pipelines/{uuid.uuid4()}/staging/meta").status_code == 404
