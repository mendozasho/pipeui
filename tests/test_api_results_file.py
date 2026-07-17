"""Behavioral guarantees for the results-report file downloads (#152, §13).

GET /pipelines/{source_id}/export/results/file?format=csv|xlsx
GET /validations/{function_id}/export/results/file?format=csv|xlsx
GET /pipelines/sets/{set_id}/export/results/file?format=csv|xlsx

  1. Each entry point streams the per-function report as an attachment named
     {label}_{date}_validation.{ext} — one row per function × source, with
     function_name and source_name populated on every row.
  2. xlsx downloads round-trip through openpyxl with the same column set.
  3. 404 for an unknown source/function/set id; 422 for a malformed id or an
     unsupported format.
  4. The temp file is removed after a successful download.
"""
from __future__ import annotations

import csv
import io
import tempfile
import uuid

import pytest
from fastapi import FastAPI
from fastapi.testclient import TestClient

from pipeui.middleware.deps import get_conn
from pipeui.middleware.pipelines import router as pipelines_router
from pipeui.middleware.validations import router as validations_router

from tests.test_api_validations import (
    get_column_id,
    register_and_ingest,
    seed_validation_fn_and_attach,
    write_fn,
)


@pytest.fixture
def client(db):
    app = FastAPI()
    app.include_router(pipelines_router)
    app.include_router(validations_router)
    app.dependency_overrides[get_conn] = lambda: db
    yield TestClient(app)


def _seed(db, tmp_path, source_name="filesrc", fn_name="pos_file"):
    source_id, _ = register_and_ingest(db, tmp_path, name=source_name)
    col_id = get_column_id(db, source_id)
    module_path = write_fn(tmp_path, fn_name, "data", "return data > 0")
    fn_id, set_id, _ = seed_validation_fn_and_attach(db, source_id, col_id, fn_name, module_path)
    return source_id, fn_id, set_id


def _csv_rows(resp):
    return list(csv.DictReader(io.StringIO(resp.text)))


# ---------------------------------------------------------------------------
# Guarantee 1: attachments with function × source rows, per entry point
# ---------------------------------------------------------------------------

@pytest.mark.integration
def test_source_results_file_csv(client, db, tmp_path):
    source_id, _, _ = _seed(db, tmp_path)

    resp = client.get(f"/pipelines/{source_id}/export/results/file?format=csv")
    assert resp.status_code == 200, resp.text
    disposition = resp.headers["content-disposition"]
    assert "attachment" in disposition
    assert "filesrc_" in disposition
    assert "_validation.csv" in disposition
    assert resp.headers["content-type"].startswith("text/csv")

    rows = _csv_rows(resp)
    assert len(rows) == 1
    assert rows[0]["function_name"] == "pos_file"
    assert rows[0]["source_name"] == "filesrc"
    assert rows[0]["status"] == "ok"


@pytest.mark.integration
def test_function_results_file_csv(client, db, tmp_path):
    _, fn_id, _ = _seed(db, tmp_path, source_name="fnsrc", fn_name="fn_route")

    resp = client.get(f"/validations/{fn_id}/export/results/file?format=csv")
    assert resp.status_code == 200, resp.text
    disposition = resp.headers["content-disposition"]
    assert "attachment" in disposition
    assert "fn_route_" in disposition
    assert "_validation.csv" in disposition

    rows = _csv_rows(resp)
    assert len(rows) == 1
    # #152 regression: the cross-source path used to export function_name=None.
    assert rows[0]["function_name"] == "fn_route"
    assert rows[0]["source_name"] == "fnsrc"


@pytest.mark.integration
def test_set_results_file_csv(client, db, tmp_path):
    _, _, set_id = _seed(db, tmp_path, source_name="setsrc", fn_name="set_route")

    resp = client.get(f"/pipelines/sets/{set_id}/export/results/file?format=csv")
    assert resp.status_code == 200, resp.text
    disposition = resp.headers["content-disposition"]
    assert "attachment" in disposition
    assert "_validation.csv" in disposition

    rows = _csv_rows(resp)
    assert len(rows) == 1
    assert rows[0]["function_name"] == "set_route"
    assert rows[0]["source_name"] == "setsrc"


# ---------------------------------------------------------------------------
# Guarantee 2: xlsx round-trip
# ---------------------------------------------------------------------------

@pytest.mark.integration
def test_source_results_file_xlsx_round_trips(client, db, tmp_path):
    from openpyxl import load_workbook

    source_id, _, _ = _seed(db, tmp_path, source_name="xlsxsrc", fn_name="xlsx_chk")

    resp = client.get(f"/pipelines/{source_id}/export/results/file?format=xlsx")
    assert resp.status_code == 200, resp.text
    assert "_validation.xlsx" in resp.headers["content-disposition"]

    ws = load_workbook(io.BytesIO(resp.content)).active
    grid = [[c.value for c in row] for row in ws.iter_rows()]
    header = {name: i for i, name in enumerate(grid[0])}
    assert "function_name" in header and "source_name" in header
    assert grid[1][header["function_name"]] == "xlsx_chk"
    assert grid[1][header["source_name"]] == "xlsxsrc"


# ---------------------------------------------------------------------------
# Guarantee 3: error contracts
# ---------------------------------------------------------------------------

def test_unknown_ids_return_404(client):
    assert client.get(f"/pipelines/{uuid.uuid4()}/export/results/file").status_code == 404
    assert client.get(f"/validations/{uuid.uuid4()}/export/results/file").status_code == 404
    assert client.get(f"/pipelines/sets/{uuid.uuid4()}/export/results/file").status_code == 404


def test_malformed_ids_return_422(client):
    assert client.get("/pipelines/not-a-uuid/export/results/file").status_code == 422
    assert client.get("/validations/not-a-uuid/export/results/file").status_code == 422
    assert client.get("/pipelines/sets/not-a-uuid/export/results/file").status_code == 422


@pytest.mark.integration
def test_unsupported_format_returns_422(client, db, tmp_path):
    source_id, _, _ = _seed(db, tmp_path, source_name="fmtsrc", fn_name="fmt_chk")
    resp = client.get(f"/pipelines/{source_id}/export/results/file?format=pdf")
    assert resp.status_code == 422


# ---------------------------------------------------------------------------
# Guarantee 4: temp-file cleanup
# ---------------------------------------------------------------------------

@pytest.mark.integration
def test_temp_file_removed_after_download(client, db, tmp_path, monkeypatch):
    import os

    created: list[str] = []
    real_mkstemp = tempfile.mkstemp

    def spying_mkstemp(*args, **kwargs):
        fd, path = real_mkstemp(*args, **kwargs)
        created.append(path)
        return fd, path

    monkeypatch.setattr(tempfile, "mkstemp", spying_mkstemp)

    source_id, _, _ = _seed(db, tmp_path, source_name="cleansrc", fn_name="clean_chk")
    resp = client.get(f"/pipelines/{source_id}/export/results/file?format=csv")
    assert resp.status_code == 200
    assert created, "expected the route to allocate a temp file"
    # TestClient runs the BackgroundTask before returning, so the file is gone.
    assert all(not os.path.exists(p) for p in created)
