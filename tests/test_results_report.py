"""Behavioral guarantees for the results-report builder and file writers (#152, §13).

build_results_report(run_result, *, source_name=None)
  1. One row per function run (function × source) from every runner output shape:
     source-tied ({steps}), cross-source ({sources} flat), set ({sources} with
     nested steps, flattened).
  2. The report columns are exactly _RESULT_COLUMNS — function identity
     (function_name, function_type) and source_name are explicit columns; the
     retired label column is absent.
  3. Source-tied steps carry no source_name — the caller-supplied kwarg fills it.
  4. A set-run source that crashed before running any function still yields one
     failed row (complete-record contract).
  5. None/empty run results produce an empty rows list, never an error.

write_results_csv / write_results_xlsx(report, dest_path)
  6. Round-trip the column set and rows; None cells become empty fields;
     pass_rate is formatted as a percentage string.
"""
from __future__ import annotations

import csv

from pipeui.backend.domain.runner.export import (
    _RESULT_COLUMNS,
    build_results_report,
    write_results_csv,
    write_results_xlsx,
)


def _step(fn="chk", status="ok", passed=2, failed=1, source_name=None):
    total = (passed or 0) + (failed or 0)
    return {
        "function_id": "abc",
        "function_name": fn,
        "function_type": "validation",
        "source_name": source_name,
        "status": status,
        "rows_passed": passed,
        "rows_failed": failed,
        "pass_rate": (passed / total) if passed is not None and total else None,
        "error": None,
        "result_id": "deadbeef",
        "label": fn,
        "failing_rows": [{"id": "r1"}],
    }


# ---------------------------------------------------------------------------
# Guarantees 1–3: the three runner shapes collapse to function × source rows
# ---------------------------------------------------------------------------

def test_source_tied_shape_stamps_caller_source_name():
    result = {"run_type": "validations", "steps": [_step("a"), _step("b")]}
    report = build_results_report(result, source_name="sales")
    assert report["columns"] == _RESULT_COLUMNS
    assert [r["function_name"] for r in report["rows"]] == ["a", "b"]
    assert all(r["source_name"] == "sales" for r in report["rows"])
    assert all(r["function_type"] == "validation" for r in report["rows"])


def test_cross_source_shape_preserves_function_identity_per_source():
    # The cross-source entries are flat (function, source) RunResults — the #152
    # regression was these rows losing function_name/function_type entirely.
    result = {
        "function_id": "abc",
        "function_name": "chk",
        "sources": [
            {**_step("chk", source_name="src1"), "source_id": "s1"},
            {**_step("chk", source_name="src2"), "source_id": "s2"},
        ],
    }
    report = build_results_report(result)
    assert len(report["rows"]) == 2
    assert {r["source_name"] for r in report["rows"]} == {"src1", "src2"}
    assert all(r["function_name"] == "chk" for r in report["rows"])


def test_set_shape_flattens_sources_by_steps():
    result = {
        "set_id": "sid",
        "set_name": "nightly",
        "sources": [
            {"source_id": "s1", "source_name": "src1",
             "steps": [_step("a"), _step("b")], "error": None},
            {"source_id": "s2", "source_name": "src2",
             "steps": [_step("a")], "error": None},
        ],
    }
    report = build_results_report(result)
    rows = report["rows"]
    assert len(rows) == 3  # 2 functions on src1 + 1 on src2
    assert [(r["function_name"], r["source_name"]) for r in rows] == [
        ("a", "src1"), ("b", "src1"), ("a", "src2"),
    ]


def test_report_columns_have_no_label():
    assert "label" not in _RESULT_COLUMNS
    for col in ("function_name", "function_type", "source_name"):
        assert col in _RESULT_COLUMNS


# ---------------------------------------------------------------------------
# Guarantee 4: a crashed set source still yields a failed row
# ---------------------------------------------------------------------------

def test_set_source_crash_with_no_steps_yields_failed_row():
    result = {
        "set_id": "sid",
        "set_name": "nightly",
        "sources": [
            {"source_id": "s1", "source_name": "src1", "steps": [], "error": "boom"},
        ],
    }
    rows = build_results_report(result)["rows"]
    assert len(rows) == 1
    assert rows[0]["status"] == "failed"
    assert rows[0]["error"] == "boom"
    assert rows[0]["source_name"] == "src1"
    assert rows[0]["rows_passed"] is None


# ---------------------------------------------------------------------------
# Guarantee 5: empty inputs
# ---------------------------------------------------------------------------

def test_none_and_empty_results_produce_empty_reports():
    for empty in (None, {}, {"steps": []}, {"sources": []}, {"steps": None}):
        report = build_results_report(empty)
        assert report == {"columns": list(_RESULT_COLUMNS), "rows": []}


# ---------------------------------------------------------------------------
# Guarantee 6: file writers
# ---------------------------------------------------------------------------

def _report():
    crashed = _step("broken", status="failed", passed=None, failed=None, source_name="src1")
    crashed["pass_rate"] = None
    crashed["error"] = "worker died"
    return build_results_report({"steps": [_step("chk", source_name="src1"), crashed]})


def test_write_results_csv_round_trips_columns_and_formats(tmp_path):
    dest = str(tmp_path / "report.csv")
    n = write_results_csv(_report(), dest)
    assert n == 2
    with open(dest, newline="") as f:
        rows = list(csv.DictReader(f))
    assert list(rows[0].keys()) == _RESULT_COLUMNS
    assert rows[0]["function_name"] == "chk"
    assert rows[0]["pass_rate"] == "66.7%"       # 2/3 formatted as a percentage
    assert rows[1]["function_name"] == "broken"  # crashed run still present
    assert rows[1]["rows_passed"] == ""          # None -> empty field
    assert rows[1]["error"] == "worker died"


def test_write_results_xlsx_round_trips_columns_and_formats(tmp_path):
    from openpyxl import load_workbook

    dest = str(tmp_path / "report.xlsx")
    n = write_results_xlsx(_report(), dest)
    assert n == 2
    ws = load_workbook(dest).active
    grid = [[c.value for c in row] for row in ws.iter_rows()]
    assert grid[0] == _RESULT_COLUMNS
    header = {name: i for i, name in enumerate(grid[0])}
    assert grid[1][header["function_name"]] == "chk"
    assert grid[1][header["pass_rate"]] == "66.7%"
    assert grid[2][header["error"]] == "worker died"
    assert grid[2][header["rows_passed"]] is None
